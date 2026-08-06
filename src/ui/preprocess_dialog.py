"""预处理交互界面：列名映射、高级设置、报告预览与导出。"""

from __future__ import annotations

from datetime import datetime
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl.styles import PatternFill
from PySide6.QtCore import QModelIndex, Qt
from PySide6.QtGui import QAction, QColor, QFontMetrics
from PySide6.QtWidgets import (
    QAbstractItemView,
    QApplication,
    QComboBox,
    QDialog,
    QDoubleSpinBox,
    QFileDialog,
    QFormLayout,
    QGroupBox,
    QHBoxLayout,
    QLabel,
    QMenu,
    QMessageBox,
    QPushButton,
    QSpinBox,
    QTableWidget,
    QTableWidgetItem,
    QVBoxLayout,
)

from src.core.preprocess import (
    DEFAULT_DATA_START_ROW,
    DEFAULT_GROUP_HEADER_ROW,
    DEFAULT_MAX_INTERPOLATION_HOURS,
    DEFAULT_MISSING_RATE_THRESHOLD,
    DEFAULT_SUB_HEADER_ROW,
    DEFAULT_ZSCORE_THRESHOLD,
    PreprocessResult,
    PreprocessSettings,
)

OUTLIER_COLOR = QColor("#FFF2CC")
INTERPOLATED_COLOR = QColor("#DDEBF7")
MISSING_COLOR = QColor("#F8CBAD")


def order_features(features: list[str]) -> list[str]:
    """模型特征顺序：进水指标在前，出水指标在后，其余（时间特征）最后。"""
    inlets = [f for f in features if f.startswith("进水")]
    outlets = [f for f in features if f.startswith("出水")]
    others = [f for f in features if f not in inlets and f not in outlets]
    return inlets + outlets + others


def format_value(value) -> str:
    """数值最多保留 3 位小数，少于 3 位保持不变；缺失值显示为空。"""
    if pd.isna(value):
        return ""
    if isinstance(value, (pd.Timestamp, np.datetime64)):
        return pd.Timestamp(value).strftime("%Y-%m-%d %H:%M:%S")
    if isinstance(value, (float, np.floating)):
        return repr(round(float(value), 3))
    return str(value)


class NoWheelComboBox(QComboBox):
    """禁止鼠标滚轮直接改选项，必须点击下拉栏选择。"""

    def wheelEvent(self, event):
        event.ignore()


class MappingTable(QTableWidget):
    """模型特征列：左键再次点击取消选中，右键可复制。"""

    def mousePressEvent(self, event):
        if event.button() == Qt.LeftButton:
            index = self.indexAt(event.pos())
            if index.isValid() and index in self.selectedIndexes():
                self.clearSelection()
                self.setCurrentIndex(QModelIndex())
                return
        super().mousePressEvent(event)


class PreprocessSetupDialog(QDialog):
    """第一步：确认列名映射、清洗阈值和高级设置。"""

    def __init__(
        self,
        path: str,
        source_type: str,
        columns: list[str],
        features: list[str],
        suggestions: dict[str, str | None],
        previous_settings: PreprocessSettings | None = None,
        parent=None,
    ):
        super().__init__(parent)
        self.setWindowTitle("数据预处理 - 列名映射与清洗设置")
        self.resize(720, 620)
        self._features = list(features)
        self._combos: dict[str, QComboBox] = {}
        display_features = order_features(features)

        type_text = {
            "unified": "统一格式表",
            "raw_file": "原始进口/出口 Excel",
            "raw_folder": "进口/出口 Excel 文件夹",
        }.get(source_type, source_type)
        path_label = QLabel(f"数据源：{path}")
        path_label.setWordWrap(True)
        type_label = QLabel(f"识别类型：{type_text}")

        layout = QVBoxLayout(self)
        layout.addWidget(path_label)
        layout.addWidget(type_label)

        mapping_label = QLabel("列名映射（模型特征 → 原始数据列）")
        mapping_label.setStyleSheet("font-weight: bold;")
        layout.addWidget(mapping_label)

        self.mapping_table = MappingTable(len(display_features), 2)
        self.mapping_table.setHorizontalHeaderLabels(["模型特征", "对应原始列"])
        self.mapping_table.verticalHeader().setVisible(False)
        self.mapping_table.setEditTriggers(QTableWidget.NoEditTriggers)
        self.mapping_table.setSelectionMode(QAbstractItemView.SingleSelection)
        self.mapping_table.setSelectionBehavior(QAbstractItemView.SelectItems)
        self.mapping_table.setContextMenuPolicy(Qt.CustomContextMenu)
        self.mapping_table.customContextMenuRequested.connect(self._copy_feature_cell)
        for i, feature in enumerate(display_features):
            self.mapping_table.setItem(i, 0, QTableWidgetItem(feature))
            combo = NoWheelComboBox()
            combo.addItem("（不映射）")
            combo.addItems(columns)
            previous = (
                previous_settings.mapping.get(feature)
                if previous_settings
                else None
            )
            initial = previous if previous in columns else suggestions.get(feature)
            if initial:
                idx = combo.findText(initial)
                combo.setCurrentIndex(idx if idx >= 0 else 0)
            self.mapping_table.setCellWidget(i, 1, combo)
            self._combos[feature] = combo
        self.mapping_table.horizontalHeader().setStretchLastSection(True)
        font_metrics = QFontMetrics(self.mapping_table.font())
        feature_width = max(font_metrics.horizontalAdvance(f) for f in features) + 32
        self.mapping_table.setColumnWidth(0, feature_width)
        layout.addWidget(self.mapping_table, 1)

        advanced = QGroupBox("高级设置（默认与训练端一致）")
        form = QFormLayout(advanced)
        self.z_spin = QDoubleSpinBox()
        self.z_spin.setRange(0.5, 20.0)
        self.z_spin.setDecimals(1)
        self.z_spin.setValue(
            previous_settings.z_score_threshold
            if previous_settings
            else DEFAULT_ZSCORE_THRESHOLD
        )
        self.gap_spin = QSpinBox()
        self.gap_spin.setRange(0, 72)
        self.gap_spin.setValue(
            previous_settings.max_interpolation_hours
            if previous_settings
            else DEFAULT_MAX_INTERPOLATION_HOURS
        )
        self.missing_spin = QDoubleSpinBox()
        self.missing_spin.setRange(0.0, 1.0)
        self.missing_spin.setDecimals(2)
        self.missing_spin.setValue(
            previous_settings.missing_rate_threshold
            if previous_settings
            else DEFAULT_MISSING_RATE_THRESHOLD
        )
        self.duplicate_combo = QComboBox()
        self.duplicate_combo.addItems(["保留第一条", "保留最后一条"])
        if previous_settings and previous_settings.duplicate_strategy == "last":
            self.duplicate_combo.setCurrentIndex(1)
        self.group_spin = QSpinBox()
        self.group_spin.setRange(0, 10)
        self.group_spin.setValue(
            previous_settings.group_header_row
            if previous_settings
            else DEFAULT_GROUP_HEADER_ROW
        )
        self.sub_spin = QSpinBox()
        self.sub_spin.setRange(0, 10)
        self.sub_spin.setValue(
            previous_settings.sub_header_row
            if previous_settings
            else DEFAULT_SUB_HEADER_ROW
        )
        self.data_spin = QSpinBox()
        self.data_spin.setRange(0, 50)
        self.data_spin.setValue(
            previous_settings.data_start_row
            if previous_settings
            else DEFAULT_DATA_START_ROW
        )
        form.addRow("Z-score 阈值", self.z_spin)
        form.addRow("最长插值小时数", self.gap_spin)
        form.addRow("高缺失率阈值", self.missing_spin)
        form.addRow("重复时间处理", self.duplicate_combo)
        form.addRow("分组表头行", self.group_spin)
        form.addRow("子表头行", self.sub_spin)
        form.addRow("数据起始行", self.data_spin)
        reset_btn = QPushButton("恢复默认")
        reset_btn.clicked.connect(self._reset_defaults)
        form.addRow("", reset_btn)
        layout.addWidget(advanced)

        buttons = QHBoxLayout()
        ok_btn = QPushButton("开始预处理")
        ok_btn.setObjectName("SecondaryButton")
        cancel_btn = QPushButton("取消")
        ok_btn.clicked.connect(self.accept)
        cancel_btn.clicked.connect(self.reject)
        buttons.addStretch(1)
        buttons.addWidget(ok_btn)
        buttons.addWidget(cancel_btn)
        layout.addLayout(buttons)

    def _reset_defaults(self):
        self.z_spin.setValue(DEFAULT_ZSCORE_THRESHOLD)
        self.gap_spin.setValue(DEFAULT_MAX_INTERPOLATION_HOURS)
        self.missing_spin.setValue(DEFAULT_MISSING_RATE_THRESHOLD)
        self.group_spin.setValue(DEFAULT_GROUP_HEADER_ROW)
        self.sub_spin.setValue(DEFAULT_SUB_HEADER_ROW)
        self.data_spin.setValue(DEFAULT_DATA_START_ROW)

    def settings(self) -> PreprocessSettings:
        mapping = {}
        for feature, combo in self._combos.items():
            text = combo.currentText()
            mapping[feature] = None if text == "（不映射）" else text
        return PreprocessSettings(
            mapping=mapping,
            z_score_threshold=float(self.z_spin.value()),
            max_interpolation_hours=int(self.gap_spin.value()),
            missing_rate_threshold=float(self.missing_spin.value()),
            duplicate_strategy="first" if self.duplicate_combo.currentText() == "保留第一条" else "last",
            group_header_row=int(self.group_spin.value()),
            sub_header_row=int(self.sub_spin.value()),
            data_start_row=int(self.data_spin.value()),
        )

    def _copy_feature_cell(self, pos):
        index = self.mapping_table.indexAt(pos)
        if index.isValid():
            self.mapping_table.setCurrentCell(index.row(), index.column())
        menu = QMenu(self)
        action = QAction("复制", menu)
        action.triggered.connect(self._copy_selected_feature)
        menu.addAction(action)
        menu.exec(self.mapping_table.viewport().mapToGlobal(pos))

    def _copy_selected_feature(self):
        item = self.mapping_table.currentItem()
        if item is not None:
            QApplication.clipboard().setText(item.text())


class PreprocessReportDialog(QDialog):
    """第二步：展示预处理报告、带颜色预览表，可导出。"""

    def __init__(self, path: str, result: PreprocessResult, parent=None):
        super().__init__(parent)
        self.setWindowTitle("预处理报告")
        self.resize(980, 680)
        self._result = result
        self._path = path

        layout = QVBoxLayout(self)
        report = QLabel(self._report_text(result.report))
        report.setWordWrap(True)
        report.setTextInteractionFlags(
            report.textInteractionFlags() | Qt.TextSelectableByMouse
        )
        layout.addWidget(report)

        preview_label = QLabel("清洗后数据预览（前 1000 行，黄色=异常值，浅蓝=插补值，红色=仍缺失）")
        preview_label.setStyleSheet("font-weight: bold;")
        layout.addWidget(preview_label)

        self.table = QTableWidget()
        self.table.setEditTriggers(QTableWidget.NoEditTriggers)
        self.table.setSelectionBehavior(QAbstractItemView.SelectItems)
        self._fill_preview()
        layout.addWidget(self.table, 1)

        export_btn = QPushButton("导出预处理结果 (.xlsx)")
        export_btn.setObjectName("SecondaryButton")
        export_btn.clicked.connect(self._export)
        ok_btn = QPushButton("确认使用并开始预测")
        back_btn = QPushButton("返回调整")
        cancel_btn = QPushButton("取消")
        back_btn.clicked.connect(self.reject)
        cancel_btn.clicked.connect(self.reject)
        ok_btn.clicked.connect(self.accept)

        buttons = QHBoxLayout()
        buttons.addWidget(export_btn)
        buttons.addStretch(1)
        buttons.addWidget(back_btn)
        buttons.addWidget(ok_btn)
        buttons.addWidget(cancel_btn)
        layout.addLayout(buttons)

    def _report_text(self, report) -> str:
        outlier_total = sum(report.outlier_counts.values())
        interpolated_total = sum(report.interpolated_counts.values())
        lines = [
            f"文件类型：{report.source_type}",
            f"原始行数：{report.original_rows} → 清洗后行数：{report.final_rows}",
            f"时间范围：{report.time_start} ~ {report.time_end}",
            f"重复时间行：{report.duplicate_rows}",
            f"异常值清除：{outlier_total} 条",
            f"缺失值插补：{interpolated_total} 条",
        ]
        if report.high_missing_features:
            lines.append("高缺失率特征：" + "、".join(report.high_missing_features))
        return "\n".join(lines)

    def _fill_preview(self):
        df = self._result.df.head(1000)
        columns = list(df.columns)
        self.table.setColumnCount(len(columns))
        self.table.setHorizontalHeaderLabels(columns)
        self.table.setRowCount(len(df))
        for r, (_, row) in enumerate(df.iterrows()):
            for c, col in enumerate(columns):
                value = row[col]
                text = format_value(value)
                item = QTableWidgetItem(text)
                if pd.isna(value):
                    item.setBackground(MISSING_COLOR)
                elif col in self._result.outlier_mask.columns and bool(
                    self._result.outlier_mask.iloc[r][col]
                ):
                    item.setBackground(OUTLIER_COLOR)
                elif col in self._result.interpolated_mask.columns and bool(
                    self._result.interpolated_mask.iloc[r][col]
                ):
                    item.setBackground(INTERPOLATED_COLOR)
                self.table.setItem(r, c, item)
        font_metrics = QFontMetrics(self.table.font())
        for c, col in enumerate(columns):
            width = font_metrics.horizontalAdvance(str(col)) + 28
            for r in range(len(df)):
                item = self.table.item(r, c)
                if item is not None:
                    width = max(width, font_metrics.horizontalAdvance(item.text()) + 24)
            self.table.setColumnWidth(c, width)

    def _export(self):
        default_name = str(Path(self._path).stem) + "_预处理结果.xlsx"
        path, _ = QFileDialog.getSaveFileName(
            self, "导出预处理结果", default_name, "Excel 文件 (*.xlsx)"
        )
        if not path:
            return
        try:
            export_preprocessed_xlsx(path, self._result)
        except PermissionError:
            alt = str(
                Path(path).with_name(
                    f"{Path(path).stem}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
                )
            )
            try:
                export_preprocessed_xlsx(alt, self._result)
            except Exception as exc:
                QMessageBox.critical(self, "导出失败", str(exc))
                return
            QMessageBox.information(
                self,
                "导出完成",
                f"原文件可能正被其他程序占用，已自动另存为：\n{alt}",
            )
            return
        except Exception as exc:
            QMessageBox.critical(self, "导出失败", str(exc))
            return
        QMessageBox.information(self, "导出完成", f"已保存：{path}")


def export_preprocessed_xlsx(path: str, result: PreprocessResult) -> None:
    """导出带颜色标记的预处理结果 Excel。"""
    df = result.df
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="预处理数据")
        ws = writer.sheets["预处理数据"]
        yellow = PatternFill("solid", fgColor="FFF2CC")
        light_blue = PatternFill("solid", fgColor="DDEBF7")
        red = PatternFill("solid", fgColor="F8CBAD")
        for col_idx, col in enumerate(df.columns, start=1):
            for row_idx in range(len(df)):
                cell = ws.cell(row=row_idx + 2, column=col_idx)
                value = df.iloc[row_idx][col]
                if pd.isna(value):
                    cell.fill = red
                elif col in result.outlier_mask.columns and bool(
                    result.outlier_mask.iloc[row_idx][col]
                ):
                    cell.fill = yellow
                elif col in result.interpolated_mask.columns and bool(
                    result.interpolated_mask.iloc[row_idx][col]
                ):
                    cell.fill = light_blue
